#!/usr/bin/env python3

from pathlib import Path
import argparse
import re
import pandas as pd
from ipaddress import ip_network
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from html import escape

rows = []

def print_banner():
    banner = r"""
     ██╗██╗   ██╗███╗   ██╗ ██████╗ ██╗   ██╗███████╗██████╗ ███████╗███████╗
     ██║██║   ██║████╗  ██║██╔═══██╗██║   ██║██╔════╝██╔══██╗██╔════╝██╔════╝
     ██║██║   ██║██╔██╗ ██║██║   ██║██║   ██║█████╗  ██████╔╝███████╗█████╗
██   ██║██║   ██║██║╚██╗██║██║   ██║╚██╗ ██╔╝██╔══╝  ██╔══██╗╚════██║██╔══╝
╚█████╔╝╚██████╔╝██║ ╚████║╚██████╔╝ ╚████╔╝ ███████╗██║  ██║███████║███████╗
 ╚════╝  ╚═════╝ ╚═╝  ╚═══╝ ╚═════╝   ╚═══╝  ╚══════╝╚═╝  ╚═╝╚══════╝╚══════╝

  Created with <3 by @nickvourd
  v1.1
"""
    print(banner)


def clean(value):
    if not value:
        return ""

    return (
        value.strip()
        .strip(";")
        .strip('"')
        .strip("'")
    )


def get_hostname(content, file_path):

    match = re.search(
        r'host-name\s+([^;\n]+);',
        content
    )

    if match:
        return clean(match.group(1))

    return file_path.stem.split("_")[0]


def cidr_from_ip(ip):

    try:
        return str(
            ip_network(ip, strict=False)
        )

    except ValueError:
        return ""


def add_row(
    hostname,
    source,
    interface,
    unit,
    subnet_name,
    ip_address,
    subnet
):

    rows.append({
        "Hostname": hostname,
        "Source": source,
        "Interface": interface,
        "Unit": unit,
        "Subnet Name": subnet_name,
        "Interface IP": ip_address,
        "Subnet": subnet
    })


# -------------------------------------------------------------------
# v1.1 fix: find ALL '{name} { ... }' blocks anywhere in the config,
# not just the first one. Real JUNOS configs from enterprise gear
# routinely place an 'interfaces { ... }' template inside groups,
# logical-systems and routing-instances BEFORE the top-level
# 'interfaces { ... }' block. The old extract_named_block() locked
# onto the first match and silently dropped every subsequent
# interface in the file.
# -------------------------------------------------------------------
def find_all_named_blocks(content, name):

    results = []

    for match in re.finditer(
        rf'\b{name}\s*\{{',
        content
    ):

        start = match.end()

        depth = 1

        i = start

        while i < len(content):

            ch = content[i]

            if ch == "{":
                depth += 1

            elif ch == "}":

                depth -= 1

                if depth == 0:

                    results.append(
                        content[start:i]
                    )

                    break

            i += 1

    return results


# kept for backward compatibility — returns the FIRST block only
def extract_named_block(content, name):

    blocks = find_all_named_blocks(
        content,
        name
    )

    return blocks[0] if blocks else ""


# -------------------------------------------------------------------
# v1.1 fix: walk only the DIRECT children of a block, instead of
# running re.finditer() across the whole text. The old code matched
# every '<id> {' in the string, including ones nested deep inside,
# which produced spurious "interfaces" like the literal '172' or
# 'inet'. This walker is brace-aware and depth-aware, and also
# tolerates 'inactive:' / 'protect:' / 'replace:' modifiers and
# JUNOS comments.
# -------------------------------------------------------------------
def iter_top_level_blocks(body):

    i = 0
    n = len(body)

    while i < n:

        ch = body[i]

        # whitespace / statement terminators
        if ch.isspace() or ch == ';':
            i += 1
            continue

        # # line comments
        if ch == '#':

            nl = body.find('\n', i)

            i = nl + 1 if nl != -1 else n

            continue

        # /* block comments */
        if body[i:i + 2] == '/*':

            end = body.find('*/', i + 2)

            i = end + 2 if end != -1 else n

            continue

        # optional config modifier (inactive:, protect:, replace:)
        mod = re.match(
            r'(?:inactive|protect|replace):\s+',
            body[i:]
        )

        if mod:
            i += mod.end()
            continue

        # try to match '<name> [optional-arg] {'
        header = re.match(
            r'([A-Za-z0-9_\-/.:*@]+)([^\n{;]*?)\{',
            body[i:]
        )

        if not header:

            # not a block — skip to the next ';' or newline
            nxt = re.search(r'[;\n]', body[i:])

            if not nxt:
                break

            i += nxt.end()

            continue

        name = header.group(1)
        arg = header.group(2).strip()

        # locate the matching '}'
        start = i + header.end()

        depth = 1
        j = start

        while j < n:

            cj = body[j]

            if cj == '{':
                depth += 1

            elif cj == '}':

                depth -= 1

                if depth == 0:
                    break

            j += 1

        if depth != 0:
            break

        yield name, arg, body[start:j]

        i = j + 1


def parse_set_format(content, hostname):

    descriptions = {}

    # Parse descriptions
    for match in re.finditer(
        r'^set\s+interfaces\s+(\S+)\s+unit\s+(\S+)\s+description\s+(.+)$',
        content,
        re.MULTILINE
    ):

        interface, unit, desc = match.groups()

        descriptions[
            (interface, unit)
        ] = clean(desc)

    # Parse interface addresses (inet and inet6)
    for match in re.finditer(
        r'^set\s+interfaces\s+(\S+)\s+unit\s+(\S+).*?\bfamily\s+inet6?\s+address\s+([0-9a-fA-F:.]+/\d+)',
        content,
        re.MULTILINE
    ):

        interface, unit, ip_address = match.groups()

        subnet = cidr_from_ip(ip_address)

        if subnet:

            subnet_name = descriptions.get(
                (interface, unit),
                f"{interface}.{unit}"
            )

            add_row(
                hostname,
                "interfaces",
                interface,
                unit,
                subnet_name,
                ip_address,
                subnet
            )


def parse_hierarchical_interfaces(
    content,
    hostname
):

    # v1.1: iterate over EVERY interfaces { ... } block in the file,
    # including ones nested inside groups, logical-systems and
    # routing-instances. The old version only saw the first one.
    interfaces_blocks = find_all_named_blocks(
        content,
        "interfaces"
    )

    if not interfaces_blocks:
        return

    for interfaces_block in interfaces_blocks:

        # walk only the direct children
        for if_name, if_arg, interface_block in (
            iter_top_level_blocks(interfaces_block)
        ):

            # interface-range FOO { ... } — treat FOO as the iface
            if if_name == "interface-range" and if_arg:
                interface = if_arg

            elif if_name in (
                "traceoptions",
                "apply-groups",
                "apply-macro"
            ):
                continue

            else:
                interface = if_name

            # walk units (and inactive units) inside this interface
            for sub_name, sub_arg, unit_block in (
                iter_top_level_blocks(interface_block)
            ):

                if sub_name != "unit":
                    continue

                unit = sub_arg

                desc_match = re.search(
                    r'\bdescription\s+("[^"]+"|[^;\n]+);',
                    unit_block
                )

                subnet_name = (
                    clean(desc_match.group(1))
                    if desc_match
                    else f"{interface}.{unit}"
                )

                addresses = re.findall(
                    r'\baddress\s+([0-9a-fA-F:.]+/\d+)',
                    unit_block
                )

                for ip_address in addresses:

                    subnet = cidr_from_ip(
                        ip_address
                    )

                    if subnet:

                        add_row(
                            hostname,
                            "interfaces",
                            interface,
                            unit,
                            subnet_name,
                            ip_address,
                            subnet
                        )


def parse_address_assignment_pools(
    content,
    hostname
):

    # v1.1: scan every 'access { ... }' block, not just the first
    access_blocks = find_all_named_blocks(
        content,
        "access"
    )

    for access_block in access_blocks:

        for name, arg, body in (
            iter_top_level_blocks(access_block)
        ):

            if name != "address-assignment":
                # the old config also allowed 'pool' directly under
                # 'access' on some platforms — keep that path open
                if name == "pool" and arg:
                    pool_name = clean(arg)
                    networks = re.findall(
                        r'\bnetwork\s+([0-9a-fA-F:.]+/\d+);',
                        body
                    )
                    for subnet in networks:
                        add_row(
                            hostname,
                            "dhcp-pool",
                            "",
                            "",
                            pool_name,
                            "",
                            subnet
                        )
                continue

            # walk pools inside address-assignment
            for sub_name, sub_arg, pool_body in (
                iter_top_level_blocks(body)
            ):

                if sub_name != "pool":
                    continue

                pool_name = clean(sub_arg)

                networks = re.findall(
                    r'\bnetwork\s+([0-9a-fA-F:.]+/\d+);',
                    pool_body
                )

                for subnet in networks:

                    add_row(
                        hostname,
                        "dhcp-pool",
                        "",
                        "",
                        pool_name,
                        "",
                        subnet
                    )


def parse_junos_file(file_path):

    content = file_path.read_text(
        errors="ignore"
    )

    hostname = get_hostname(
        content,
        file_path
    )

    print(f"[+] Parsing {file_path.name}")

    parse_set_format(
        content,
        hostname
    )

    parse_hierarchical_interfaces(
        content,
        hostname
    )

    parse_address_assignment_pools(
        content,
        hostname
    )


def export_excel(output_file):

    df = pd.DataFrame(rows)

    if df.empty:
        print("[-] No JUNOS subnets found.")
        return

    df = df.drop_duplicates()

    df = df.sort_values([
        "Hostname",
        "Source",
        "Interface",
        "Unit",
        "Subnet"
    ])

    df.to_excel(
        output_file,
        index=False
    )

    wb = load_workbook(output_file)

    ws = wb.active

    ws.title = "JUNOS Subnets"

    # Auto-size columns
    for column in ws.columns:

        max_length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column
        )

        ws.column_dimensions[
            get_column_letter(
                column[0].column
            )
        ].width = max_length + 4

    # Create Excel table
    table_range = (
        f"A1:G{ws.max_row}"
    )

    table = Table(
        displayName="JunosSubnetTable",
        ref=table_range
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style

    ws.add_table(table)

    ws.freeze_panes = "A2"

    wb.save(output_file)

    print(
        f"[+] Exported Excel to {output_file}"
    )


def detect_segment_type(subnet_name):

    name = subnet_name.lower()

    if (
        "printer" in name or
        "print" in name
    ):
        return "Printers", "🖨️"

    if (
        "server" in name or
        "srv" in name
    ):
        return "Servers", "🖥️"

    if (
        "voice" in name or
        "phone" in name
    ):
        return "Voice", "☎️"

    if (
        "wifi" in name or
        "wireless" in name or
        "wlan" in name
    ):
        return "WiFi", "📡"

    if "guest" in name:
        return "Guest", "👥"

    if (
        "camera" in name or
        "cctv" in name
    ):
        return "Cameras", "📷"

    if "iot" in name:
        return "IoT", "🔌"

    if (
        "mgmt" in name or
        "management" in name or
        "netinfra" in name
    ):
        return "Management", "🛠️"

    if "dmz" in name:
        return "DMZ", "🧱"

    if (
        "user" in name or
        "client" in name
    ):
        return "Users", "💻"

    return "Network", "🌐"


def export_html(output_file):

    df = pd.DataFrame(rows)

    if df.empty:
        print("[-] No HTML data found.")
        return

    df = df.drop_duplicates()

    df = df.sort_values([
        "Hostname",
        "Subnet Name",
        "Subnet"
    ])

    df["Segment Type"] = (
        df["Subnet Name"].apply(
            lambda x:
            detect_segment_type(
                str(x)
            )[0]
        )
    )

    df["Icon"] = (
        df["Subnet Name"].apply(
            lambda x:
            detect_segment_type(
                str(x)
            )[1]
        )
    )

    categories = ""

    for segment_type, group in (
        df.groupby("Segment Type")
    ):

        icon = (
            group.iloc[0]["Icon"]
        )

        cards = ""

        for _, row in group.iterrows():

            cards += f"""
            <div class="node">

                <div class="node-icon">
                    {escape(str(row["Icon"]))}
                </div>

                <div class="node-title">
                    {escape(str(row["Subnet Name"]))}
                </div>

                <div class="node-subnet">
                    {escape(str(row["Subnet"]))}
                </div>

                <div class="node-meta">
                    {escape(str(row["Hostname"]))}<br>

                    {escape(str(row["Interface"]))}.
                    {escape(str(row["Unit"]))}
                </div>

            </div>
            """

        categories += f"""
        <section class="segment-group">

            <div class="segment-header">

                <span class="segment-icon">
                    {escape(str(icon))}
                </span>

                <span>
                    {escape(str(segment_type))}
                </span>

                <span class="count">
                    {len(group)} segments
                </span>

            </div>

            <div class="nodes">
                {cards}
            </div>

        </section>
        """

    html = f"""
<!DOCTYPE html>
<html>

<head>

<meta charset="UTF-8">

<title>
Network Segmentation Map
</title>

<style>

body {{
    font-family: Arial, sans-serif;
    background: #eef2f7;
    margin: 0;
    padding: 30px;
}}

h1 {{
    text-align: center;
    color: #111827;
    margin-bottom: 5px;
}}

.subtitle {{
    text-align: center;
    color: #6b7280;
    margin-bottom: 35px;
}}

.map {{
    max-width: 1500px;
    margin: auto;
}}

.segment-group {{
    background: white;
    border-radius: 18px;
    padding: 22px;
    margin-bottom: 28px;
    box-shadow:
        0 8px 24px rgba(0,0,0,0.08);

    border:
        1px solid #e5e7eb;
}}

.segment-header {{
    display: flex;
    align-items: center;
    gap: 12px;
    font-size: 22px;
    font-weight: bold;
    color: #111827;
    margin-bottom: 20px;
    border-bottom:
        2px solid #e5e7eb;

    padding-bottom: 12px;
}}

.segment-icon {{
    font-size: 32px;
}}

.count {{
    margin-left: auto;
    font-size: 14px;
    color: #6b7280;
    background: #f3f4f6;
    padding: 6px 10px;
    border-radius: 999px;
}}

.nodes {{
    display: flex;
    flex-wrap: wrap;
    gap: 18px;
}}

.node {{
    width: 230px;
    min-height: 150px;
    background: #f9fafb;
    border:
        2px solid #d1d5db;

    border-radius: 16px;
    padding: 16px;
    text-align: center;
    position: relative;
}}

.node::before {{
    content: "";
    position: absolute;
    top: -18px;
    left: 50%;
    width: 2px;
    height: 18px;
    background: #9ca3af;
}}

.node-icon {{
    font-size: 38px;
    margin-bottom: 8px;
}}

.node-title {{
    font-weight: bold;
    color: #111827;
    margin-bottom: 8px;
    word-break: break-word;
}}

.node-subnet {{
    color: #2563eb;
    font-weight: bold;
    margin-bottom: 8px;
}}

.node-meta {{
    font-size: 12px;
    color: #6b7280;
    line-height: 1.4;
}}

.legend {{
    max-width: 1500px;
    margin: 0 auto 25px auto;
    background: white;
    border-radius: 14px;
    padding: 16px;
    display: flex;
    flex-wrap: wrap;
    gap: 14px;
    justify-content: center;
    box-shadow:
        0 4px 16px rgba(0,0,0,0.06);
}}

.legend-item {{
    background: #f3f4f6;
    border-radius: 999px;
    padding: 8px 12px;
    font-size: 14px;
}}

</style>

</head>

<body>

<h1>
Network Segmentation Map
</h1>

<div class="subtitle">

Total discovered segments:
<b>{len(df)}</b>

</div>

<div class="legend">

<div class="legend-item">
💻 Users
</div>

<div class="legend-item">
🖥️ Servers
</div>

<div class="legend-item">
🖨️ Printers
</div>

<div class="legend-item">
📡 WiFi
</div>

<div class="legend-item">
☎️ Voice
</div>

<div class="legend-item">
🛠️ Management
</div>

<div class="legend-item">
🧱 DMZ
</div>

<div class="legend-item">
🌐 Other
</div>

</div>

<div class="map">
{categories}
</div>

</body>
</html>
"""

    Path(output_file).write_text(
        html,
        encoding="utf-8"
    )

    print(
        f"[+] Exported HTML map to {output_file}"
    )


def main():

    print_banner()

    parser = argparse.ArgumentParser(
        description=(
            "Universal JUNOS subnet parser "
            "with Excel + HTML network map export"
        )
    )

    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Input .conf file or folder"
    )

    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output Excel file"
    )

    parser.add_argument(
        "--html",
        required=False,
        help="Optional HTML map output"
    )

    args = parser.parse_args()

    input_path = Path(args.input)

    if not input_path.exists():

        print(
            "[-] Input path does not exist."
        )

        return

    # Single file
    if input_path.is_file():

        parse_junos_file(
            input_path
        )

    # Folder
    elif input_path.is_dir():

        conf_files = sorted(
            input_path.glob("*.conf")
        )

        if not conf_files:

            print(
                "[-] No .conf files found."
            )

            return

        for conf_file in conf_files:

            parse_junos_file(
                conf_file
            )

    export_excel(
        args.output
    )

    if args.html:

        export_html(
            args.html
        )


if __name__ == "__main__":
    main()
